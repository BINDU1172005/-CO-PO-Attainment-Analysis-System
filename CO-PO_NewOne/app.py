from flask import Flask, render_template, request, send_file, abort
import pandas as pd
from io import BytesIO
import numpy as np
import zipfile
# openpyxl is needed for manual worksheet manipulation (merging/centering)
from openpyxl.styles import Alignment, Font 
from openpyxl.chart import BarChart, Reference, Series 
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter # Used for column width

app = Flask(__name__)

# holds the last generated Excel results in memory (per process)
last_results_excel_bytes = None


def get_co_attainment_level(percentage_of_students, thr3_pct, thr2_pct, thr1_pct):
    # ... (unchanged)
    p = percentage_of_students
    t3 = thr3_pct / 100.0
    t2 = thr2_pct / 100.0
    t1 = thr1_pct / 100.0

    if p >= t3:
        return 3
    elif p >= t2:
        return 2
    elif p >= t1:
        return 1
    else:
        return 0


def calculate_tool_co_attainment(df_marks, df_tool_map_meta, tool_type,
                                 threshold_percentage, thr3_pct, thr2_pct, thr1_pct):
    # ... (unchanged)
    df_tool_map = df_tool_map_meta[df_tool_map_meta['Tool_Question'].str.startswith(tool_type)]
    if df_tool_map.empty or df_marks.empty:
        return {}

    co_results = {}
    total_students = len(df_marks)
    all_cos = df_tool_map['CO'].unique()

    for co_name in all_cos:
        co_questions = df_tool_map[df_tool_map['CO'] == co_name]
        mark_cols_for_co = [
            q + '_Marks' for q in co_questions['Tool_Question'].tolist()
            if q + '_Marks' in df_marks.columns
        ]
        if not mark_cols_for_co:
            co_results[co_name] = 0
            continue

        max_marks = co_questions['Max_Marks'].sum()
        if max_marks == 0:
            co_results[co_name] = 0
            continue

        obtained_marks_sum = df_marks[mark_cols_for_co].fillna(0).sum(axis=1)
        threshold_score = max_marks * (threshold_percentage / 100.0)
        students_above_threshold = (obtained_marks_sum >= threshold_score).sum()
        percentage = students_above_threshold / total_students if total_students > 0 else 0
        attainment_level = get_co_attainment_level(percentage, thr3_pct, thr2_pct, thr1_pct)
        co_results[co_name] = attainment_level

    return co_results


def calculate_final_direct_co_attainment_weighted(all_tool_attainments, df_tool_map_meta,
                                                  cie_weight, see_weight):
    # ... (unchanged)
    total = float(cie_weight + see_weight) if (cie_weight + see_weight) != 0 else 1.0
    cie_w = cie_weight / total
    see_w = see_weight / total

    final_direct_co = {}
    all_cos = df_tool_map_meta['CO'].unique()
    cie_tools = df_tool_map_meta[df_tool_map_meta['Assessment_Type'] == 'CIE'][
        'Tool_Question'
    ].str.split('_', expand=True)[0].unique()
    see_tool = df_tool_map_meta[df_tool_map_meta['Assessment_Type'] == 'SEE'][
        'Tool_Question'
    ].str.split('_', expand=True)[0].unique()
    see_tool_key = see_tool[0] if len(see_tool) > 0 else None

    for co in all_cos:
        cie_attainments = []
        see_attainment = 0

        for tool_key, co_results in all_tool_attainments.items():
            if tool_key in cie_tools:
                cie_attainments.append(co_results.get(co, 0))

        if see_tool_key and see_tool_key in all_tool_attainments:
            see_attainment = all_tool_attainments[see_tool_key].get(co, 0)

        avg_cie = np.mean(cie_attainments) if cie_attainments else 0
        final_val = (cie_w * avg_cie) + (see_w * see_attainment)
        final_direct_co[co] = round(final_val, 3)

    return final_direct_co


def calculate_indirect_co_attainment(df_survey):
    # ... (unchanged)
    indirect_results = {}
    co_rating_columns = [col for col in df_survey.columns if 'CO' in col and '_Rating' in col]
    if df_survey.empty:
        return {col.replace('_Rating', ''): 0 for col in co_rating_columns}

    for col in co_rating_columns:
        avg_rating = df_survey[col].mean()
        co_name = col.replace('_Rating', '')
        indirect_results[co_name] = round(avg_rating, 3)

    return indirect_results


def calculate_po_attainment(final_co_attainments, df_mapping):
    # ... (unchanged)
    po_results = {}
    df_mapping_copy = df_mapping.copy()
    if 'CO' in df_mapping_copy.columns:
        df_mapping_copy.set_index('CO', inplace=True)

    po_pso_cols = [col for col in df_mapping_copy.columns
                   if col.startswith('PO') or col.startswith('PSO')]

    for po_col in po_pso_cols:
        numerator = 0
        denominator = 0
        for co_name, mapping_level in df_mapping_copy[po_col].items():
            try:
                mapping_level = float(mapping_level)
            except (ValueError, TypeError):
                continue
            if pd.notna(mapping_level) and mapping_level > 0 and co_name in final_co_attainments:
                numerator += final_co_attainments[co_name] * mapping_level
                denominator += mapping_level

        po_results[po_col] = round(numerator / denominator, 3) if denominator > 0 else 0

    return po_results


# --- Helper: Adds data and a chart to a sheet (Final version of helpers) ---

def setup_results_sheet(writer, df, sheet_name, cn, dn, cc):
    # ... (unchanged)
    col_count = len(df.columns)
    data_row_count = len(df)
    
    # 1. Write the DataFrame starting at row 5 (A5)
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=4) 
    
    # 2. Write metadata header (writes to rows 1 and 2)
    write_metadata_to_sheet(writer, sheet_name, cn, dn, cc, col_count)
    
    worksheet = writer.sheets[sheet_name]
    start_row_for_data = 5 # Excel Row 5 is header, Data starts at Row 6

    # --- 3. Auto-Size Columns ---
    from openpyxl.utils import get_column_letter 
    for col in range(1, col_count + 1):
        max_length = 0
        column = get_column_letter(col)
        
        max_length = max(len(str(worksheet.cell(row=start_row_for_data, column=col).value)), max_length)
        for row in range(start_row_for_data + 1, start_row_for_data + data_row_count + 1):
            length = len(str(worksheet.cell(row=row, column=col).value))
            max_length = max(max_length, length)

        worksheet.column_dimensions[column].width = max(max_length * 1.2, 10)
    
    worksheet.column_dimensions['A'].width = max(worksheet.column_dimensions['A'].width, 15)
    
    # --- 4. Add Chart (The FINAL Working Logic) ---
    if data_row_count == 0:
        return

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = f"{sheet_name.replace('_', ' ')} Attainment Levels"
    chart.y_axis.title = "Attainment Level (0-3 Scale)"
    chart.y_axis.scaling.max = 3.0
    chart.shape = 4
    
    chart_anchor = f'{chr(ord("A") + col_count + 2)}5' 

    # 1. Data Range: Attainment Level column (B5 to B[end])
    data_range = Reference(worksheet, min_col=2, min_row=start_row_for_data, 
                           max_col=2, max_row=start_row_for_data + data_row_count)
    
    # 2. Categories Range: Course Outcome column (A6 to A[end])
    cats_col_ref = Reference(worksheet, min_col=1, min_row=start_row_for_data + 1, 
                             max_col=1, max_row=start_row_for_data + data_row_count)
    
    # Add data: titles_from_data=True uses row 5 (B5) as the series title.
    chart.add_data(data_range, titles_from_data=True) 
    
    # Set categories (This provides the CO1, CO2 labels on the X-axis)
    chart.set_categories(cats_col_ref)
    
    # --- CRITICAL FINAL TWEAKS TO ENABLE VISIBILITY ---
    
    # 1. Remove the Redundant Legend
    chart.legend = None

    # 2. Enable Data Labels (THE FINAL FIX)
    if chart.series:
        # The series titles (the header 'Attainment Level') is redundant, remove it
        chart.series[0].title = None
        
        # Add the Data Labels to the primary series
        chart.series[0].dLbls = DataLabelList() 
        chart.series[0].dLbls.showVal = True    # Show the Value (2.4, 0.4, 0.6)
        chart.series[0].dLbls.showCat = True    # Show the Category Name (CO1, CO2, CO3)
        chart.series[0].dLbls.showSerName = False # Hide Series Name
        
        # Fix for X-axis labels (already correct, just for robustness)
        if chart.x_axis:
            chart.x_axis.delete = False
            chart.x_axis.tickLblPos = 'low' 
    
    # Position the chart
    worksheet.add_chart(chart, chart_anchor) 


def write_metadata_to_sheet(writer, sheet_name, cn, dn, cc, col_count):
    """Writes metadata headers to the sheet, with the left-aligned look from the screenshot."""
    worksheet = writer.sheets[sheet_name]
    
    # Define a left-aligned style for the header text
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True) 
    header_font = Font(bold=True, size=12)
    
    # 1. College Name and Department (Merged on A1:B1 for two columns)
    worksheet.merge_cells('A1:B1')
    worksheet['A1'].value = f'{cn} - {dn}'
    worksheet['A1'].alignment = left_align
    worksheet['A1'].font = header_font
    
    # 2. Course Code (Merged on A2:B2)
    worksheet.merge_cells('A2:B2')
    worksheet['A2'].value = f'Course Code: {cc}'
    worksheet['A2'].alignment = left_align
    worksheet['A2'].font = header_font
    
    # Data is expected to start at row 5 (A5), so return 5
    return 5 



def run_calculation_pipeline(
    df_marks, df_tool_map_meta, df_co_po_mapping, df_survey,
    threshold_percentage,
    thr3_pct=70, thr2_pct=55, thr1_pct=40,
    cie_weight=60, see_weight=40,
    direct_weight=0.8, indirect_weight=0.2,
    # --- METADATA ARGUMENTS ---
    cn='College Name', dn='Department Name', cc='Course Code'
):
    
    # ensure types
    df_tool_map_meta['CO'] = df_tool_map_meta['CO'].astype(str)
    tool_prefixes = df_tool_map_meta['Tool_Question'].str.split('_', expand=True)[0].unique()

    # per-tool attainment
    all_tool_attainments = {}
    for tool in tool_prefixes:
        tool_attainments = calculate_tool_co_attainment(
            df_marks, df_tool_map_meta, tool,
            threshold_percentage, thr3_pct, thr2_pct, thr1_pct
        )
        all_tool_attainments[tool] = tool_attainments

    # direct and indirect
    final_direct_co = calculate_final_direct_co_attainment_weighted(
        all_tool_attainments, df_tool_map_meta, cie_weight, see_weight
    )
    indirect_co = calculate_indirect_co_attainment(df_survey)

    # normalize direct/indirect weights
    total_di = float(direct_weight + indirect_weight) if (direct_weight + indirect_weight) != 0 else 1.0
    dw = direct_weight / total_di
    iw = indirect_weight / total_di

    final_co = {}
    all_co_keys = set(final_direct_co.keys()) | set(indirect_co.keys())
    for co in sorted(list(all_co_keys)):
        direct_val = final_direct_co.get(co, 0)
        indirect_val = indirect_co.get(co, 0)
        final_val = (dw * direct_val) + (iw * indirect_val)
        final_co[co] = round(final_val, 3)

    final_po = calculate_po_attainment(final_co, df_co_po_mapping)

    # DataFrames (Use consistent names for headers for chart setup)
    df_direct_co = pd.DataFrame(list(final_direct_co.items()),
                                columns=['Course Outcome', 'Attainment Level'])
    df_indirect_co = pd.DataFrame(list(indirect_co.items()),
                                  columns=['Course Outcome', 'Attainment Level'])
    df_final_co = pd.DataFrame(list(final_co.items()),
                               columns=['Course Outcome', 'Attainment Level'])
    df_final_po = pd.DataFrame(list(final_po.items()),
                               columns=['Program/Skill Outcome', 'Attainment Level'])

    # Build Excel in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        results_dfs = {
            'Direct_CO': df_direct_co,
            'Indirect_CO': df_indirect_co,
            'Final_CO': df_final_co,
            'Final_PO': df_final_po,
        }
        
        for sheet_name, df in results_dfs.items():
            # Call the unified setup function to write headers, data, and chart
            setup_results_sheet(writer, df, sheet_name, cn, dn, cc)
            
    output.seek(0)
    excel_bytes = output.getvalue()

    # Re-set HTML headers to be descriptive (match expected results.html headers)
    df_direct_co.columns = ['Course Outcome', 'Attainment Level (60%CIE+40%SEE)']
    df_indirect_co.columns = ['Course Outcome', 'Attainment Level (Survey Avg 1-3)']
    df_final_co.columns = ['Course Outcome', 'Final Attainment Level (80%D+20%I)']
    df_final_po.columns = ['Program/Skill Outcome', 'Attainment Level']

    return df_direct_co, df_indirect_co, df_final_co, df_final_po, excel_bytes


# --- Routes ---

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/calculate', methods=['POST'])
def calculate():
    global last_results_excel_bytes

    try:
        # basic
        threshold_percentage = float(request.form.get('threshold', 60))
        input_method = request.form.get('input_method', 'upload')

        # --- Retrieve Metadata for Excel Branding (from hidden inputs) ---
        college_name = request.form.get('college_name_calc', 'College Name')
        dept_name = request.form.get('department_name_calc', 'Department Name')
        course_code = request.form.get('course_code_calc', 'Course Code')
        
        # configuration inputs
        thr3_pct = float(request.form.get('level3_pct', 70))
        thr2_pct = float(request.form.get('level2_pct', 55))
        thr1_pct = float(request.form.get('level1_pct', 40))

        cie_weight = float(request.form.get('cie_weight', 60))
        see_weight = float(request.form.get('see_weight', 40))

        direct_weight = float(request.form.get('direct_weight', 0.8))
        indirect_weight = float(request.form.get('indirect_weight', 0.2))

        # load files
        if input_method == 'upload':
            files = request.files
            required_files = ['all_data_file', 'co_po_mapping_file', 'survey_file']
            if not all(f in files and files[f].filename for f in required_files):
                return render_template('error.html',
                                       error="Please upload all three required files: Student Data, CO-PO Mapping, and Survey.")
            all_sheets = pd.read_excel(files['all_data_file'], sheet_name=None)
            df_marks = all_sheets.get('1_Student_Marks')
            df_tool_map_meta = all_sheets.get('2_Tool_CO_Mapping')
            if df_marks is None or df_tool_map_meta is None:
                raise ValueError("Student Data must have sheets '1_Student_Marks' and '2_Tool_CO_Mapping'.")
            df_co_po_mapping = pd.read_excel(files['co_po_mapping_file'])
            df_survey = pd.read_excel(files['survey_file'])

        else:  # link
            urls = request.form
            required_urls = ['all_data_url', 'co_po_mapping_url', 'survey_url']
            if not all(urls.get(u) for u in required_urls):
                return render_template('error.html',
                                       error="Please provide all three required URLs: Student Data, CO-PO Mapping, and Survey.")
            all_sheets = pd.read_excel(urls['all_data_url'], sheet_name=None)
            df_marks = all_sheets.get('1_Student_Marks')
            df_tool_map_meta = all_sheets.get('2_Tool_CO_Mapping')
            if df_marks is None or df_tool_map_meta is None:
                raise ValueError("Student Data URL must have sheets '1_Student_Marks' and '2_Tool_CO_Mapping'.")
            df_co_po_mapping = pd.read_excel(urls['co_po_mapping_url'])
            df_survey = pd.read_excel(urls['survey_url'])

        # run pipeline
        (df_direct_co, df_indirect_co,
         df_final_co, df_final_po,
         excel_bytes) = run_calculation_pipeline(
            df_marks, df_tool_map_meta, df_co_po_mapping, df_survey,
            threshold_percentage,
            thr3_pct, thr2_pct, thr1_pct,
            cie_weight, see_weight,
            direct_weight, indirect_weight,
            # Pass metadata to the pipeline
            cn=college_name, dn=dept_name, cc=course_code 
        )

        # store Excel bytes for download_results
        last_results_excel_bytes = excel_bytes

        # Re-set HTML headers to be descriptive (match expected results.html headers)
        df_direct_co.columns = ['Course Outcome', 'Attainment Level (60%CIE+40%SEE)']
        df_indirect_co.columns = ['Course Outcome', 'Attainment Level (Survey Avg 1-3)']
        df_final_co.columns = ['Course Outcome', 'Final Attainment Level (80%D+20%I)']
        df_final_po.columns = ['Program/Skill Outcome', 'Attainment Level']

        return render_template(
            'results.html',
            direct_co_table=df_direct_co.to_html(classes='table table-hover table-sm', index=False),
            indirect_co_table=df_indirect_co.to_html(classes='table table-hover table-sm', index=False),
            final_co_table=df_final_co.to_html(classes='table table-hover table-sm', index=False),
            final_po_table=df_final_po.to_html(classes='table table-hover table-sm', index=False)
        )

    except Exception as e:
        error_message = f"A critical error occurred: {e}. Please check your files and configuration inputs."
        return render_template('error.html', error=error_message)


@app.route('/download_results', methods=['GET'])
def download_results():
    # ... (unchanged)
    global last_results_excel_bytes
    if not last_results_excel_bytes:
        return render_template('error.html',
                               error="No results available yet. Please run a calculation first.")
    buf = BytesIO(last_results_excel_bytes)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='CO_PO_Attainment_Results.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# --- Sample file generators (unchanged) ---

@app.route('/download_sample/<sample_name>')
def download_sample(sample_name):
    # ... (unchanged)
    name = sample_name.lower()
    
    # --- Retrieve metadata from query parameters ---
    course_code = request.args.get('cc', 'sample') 
    college_name = request.args.get('cn', 'Sample College')
    dept_name = request.args.get('dn', 'Sample Dept')
    
    if name == 'student':
        # --- NEW DETAILED SAMPLE DATA ---
        df_marks = pd.DataFrame({
            "USN": ["1RV01", "1RV02", "1RV03"],
            "STUDENT NAME": ["Student A", "Student B", "Student C"],
            "T1_Q1a_Marks": [6, 4, 3],
            "T1_Q1b_Marks": [4, 2, 1],
            "T1_Q2a_Marks": [5, 4, 2],
            "T1_Q2b_Marks": [5, 5, 3],
            "T2_Q1a_Marks": [6, 5, 4],
            "T2_Q1b_Marks": [4, 3, 2],
            "T2_Q2a_Marks": [5, 4, 3],
            "T2_Q2b_Marks": [5, 4, 2],
            "ASSIGN_Q1_Marks": [10, 8, 5],
            "ASSIGN_Q2_Marks": [10, 7, 4],
            "SEE_Q1a_Marks": [6, 5, 4],
            "SEE_Q1b_Marks": [4, 3, 2],
            "SEE_Q1c_Marks": [10, 8, 6],
            "SEE_Q2a_Marks": [10, 7, 5]
        })
        df_tool_map = pd.DataFrame({
            "Tool_Question": ["T1_Q1a", "T1_Q1b", "T1_Q2a", "T1_Q2b", 
                              "T2_Q1a", "T2_Q1b", "T2_Q2a", "T2_Q2b", 
                              "ASSIGN_Q1", "ASSIGN_Q2", 
                              "SEE_Q1a", "SEE_Q1b", "SEE_Q1c", "SEE_Q2a"],
            "CO": ["CO1", "CO1", "CO2", "CO2",
                   "CO1", "CO1", "CO3", "CO3",
                   "CO1", "CO2",
                   "CO1", "CO1", "CO2", "CO3"],
            "Max_Marks": [6, 4, 5, 5, 
                          6, 4, 5, 5, 
                          10, 10, 
                          6, 4, 10, 10],
            "Assessment_Type": ["CIE", "CIE", "CIE", "CIE",
                                "CIE", "CIE", "CIE", "CIE",
                                "CIE", "CIE",
                                "SEE", "SEE", "SEE", "SEE"]
        })
        # --- END NEW DETAILED SAMPLE DATA ---
        
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            # Write Student Marks Data
            sheet_marks_name = "1_Student_Marks"
            df_marks.to_excel(writer, sheet_name=sheet_marks_name, index=False, startrow=5)
            write_metadata_to_sheet(writer, sheet_marks_name, college_name, dept_name, course_code, len(df_marks.columns))
            
            # Write Tool Mapping Data
            sheet_map_name = "2_Tool_CO_Mapping"
            df_tool_map.to_excel(writer, sheet_name=sheet_map_name, index=False, startrow=5)
            write_metadata_to_sheet(writer, sheet_map_name, college_name, dept_name, course_code, len(df_tool_map.columns))
            
        buf.seek(0)
        filename = f"{course_code}_student_data_sample.xlsx"

    elif name in ('copomatrix', 'copomap'):
        # --- NEW DEFAULT: 11 POs and 2 PSOs ---
        po_names = [f'PO{i}' for i in range(1, 12)]
        pso_names = [f'PSO{i}' for i in range(1, 3)]
        all_cols = ['CO'] + po_names + pso_names
        
        # Create data with some sample mapping levels (1, 2, or 3)
        data = {
            'CO': ['CO1', 'CO2', 'CO3'],
            'PO1': [3, 1, 0], 'PO2': [1, 2, 3], 'PO3': [2, 3, 1], 'PO4': [0, 1, 2],
            'PO5': [1, 0, 3], 'PO6': [3, 2, 1], 'PO7': [2, 1, 0], 'PO8': [1, 3, 2],
            'PO9': [2, 0, 1], 'PO10': [0, 2, 3], 'PO11': [3, 1, 2],
            'PSO1': [2, 3, 1], 'PSO2': [1, 1, 2]
        }
        df_copom = pd.DataFrame(data, columns=all_cols)
        # --- END NEW DEFAULT ---

        buf = BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            sheet_name = "CO_PO_Mapping"
            df_copom.to_excel(writer, sheet_name=sheet_name, index=False, startrow=5)
            write_metadata_to_sheet(writer, sheet_name, college_name, dept_name, course_code, len(df_copom.columns))
            
        buf.seek(0)
        filename = f"{course_code}_co_po_mapping_sample.xlsx"

    elif name == 'survey':
        # NOTE: Updated survey to include CO3
        df_survey = pd.DataFrame({
            "USN": ["1RV01", "1RV02", "1RV03"],
            "CO1_Rating": [3, 2, 3],
            "CO2_Rating": [2, 1, 2],
            "CO3_Rating": [3, 2, 3]
        })
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            sheet_name = "Survey"
            df_survey.to_excel(writer, sheet_name=sheet_name, index=False, startrow=5)
            write_metadata_to_sheet(writer, sheet_name, college_name, dept_name, course_code, len(df_survey.columns))
            
        buf.seek(0)
        filename = f"{course_code}_survey_data_sample.xlsx"

    else:
        return abort(404, description="Sample not found")

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/download_sample/all')
def download_all_samples():
    # ... (unchanged)
    # --- Retrieve metadata from query parameters ---
    course_code = request.args.get('cc', 'sample')
    college_name = request.args.get('cn', 'Sample College')
    dept_name = request.args.get('dn', 'Sample Dept')
    
    mem_zip = BytesIO()
    with zipfile.ZipFile(mem_zip, mode='w', compression=zipfile.ZIP_DEFLATED) as z:
        
        # 1. Student Data (Marks + Mapping)
        sb = BytesIO()
        # --- NEW DETAILED SAMPLE DATA ---
        df_marks = pd.DataFrame({
            "USN": ["1RV01", "1RV02", "1RV03"],
            "STUDENT NAME": ["Student A", "Student B", "Student C"],
            "T1_Q1a_Marks": [6, 4, 3],
            "T1_Q1b_Marks": [4, 2, 1],
            "T1_Q2a_Marks": [5, 4, 2],
            "T1_Q2b_Marks": [5, 5, 3],
            "T2_Q1a_Marks": [6, 5, 4],
            "T2_Q1b_Marks": [4, 3, 2],
            "T2_Q2a_Marks": [5, 4, 3],
            "T2_Q2b_Marks": [5, 4, 2],
            "ASSIGN_Q1_Marks": [10, 8, 5],
            "ASSIGN_Q2_Marks": [10, 7, 4],
            "SEE_Q1a_Marks": [6, 5, 4],
            "SEE_Q1b_Marks": [4, 3, 2],
            "SEE_Q1c_Marks": [10, 8, 6],
            "SEE_Q2a_Marks": [10, 7, 5]
        })
        df_tool_map = pd.DataFrame({
            "Tool_Question": ["T1_Q1a", "T1_Q1b", "T1_Q2a", "T1_Q2b", 
                              "T2_Q1a", "T2_Q1b", "T2_Q2a", "T2_Q2b", 
                              "ASSIGN_Q1", "ASSIGN_Q2", 
                              "SEE_Q1a", "SEE_Q1b", "SEE_Q1c", "SEE_Q2a"],
            "CO": ["CO1", "CO1", "CO2", "CO2",
                   "CO1", "CO1", "CO3", "CO3",
                   "CO1", "CO2",
                   "CO1", "CO1", "CO2", "CO3"],
            "Max_Marks": [6, 4, 5, 5, 
                          6, 4, 5, 5, 
                          10, 10, 
                          6, 4, 10, 10],
            "Assessment_Type": ["CIE", "CIE", "CIE", "CIE",
                                "CIE", "CIE", "CIE", "CIE",
                                "CIE", "CIE",
                                "SEE", "SEE", "SEE", "SEE"]
        })
        # --- END NEW DETAILED SAMPLE DATA ---

        with pd.ExcelWriter(sb, engine='openpyxl') as writer:
            # Write Marks
            sheet_marks_name = "1_Student_Marks"
            df_marks.to_excel(writer, sheet_name=sheet_marks_name, index=False, startrow=5)
            write_metadata_to_sheet(writer, sheet_marks_name, college_name, dept_name, course_code, len(df_marks.columns))
            
            # Write Tool Mapping
            sheet_map_name = "2_Tool_CO_Mapping"
            df_tool_map.to_excel(writer, sheet_name=sheet_map_name, index=False, startrow=5)
            write_metadata_to_sheet(writer, sheet_map_name, college_name, dept_name, course_code, len(df_tool_map.columns))

        sb.seek(0)
        z.writestr(f'{course_code}_student_data_sample.xlsx', sb.getvalue()) 

        # 2. CO-PO Mapping
        cb = BytesIO()
        po_names = [f'PO{i}' for i in range(1, 12)]
        pso_names = [f'PSO{i}' for i in range(1, 3)]
        all_cols = ['CO'] + po_names + pso_names
        data = {
            'CO': ['CO1', 'CO2', 'CO3'],
            'PO1': [3, 1, 0], 'PO2': [1, 2, 3], 'PO3': [2, 3, 1], 'PO4': [0, 1, 2],
            'PO5': [1, 0, 3], 'PO6': [3, 2, 1], 'PO7': [2, 1, 0], 'PO8': [1, 3, 2],
            'PO9': [2, 0, 1], 'PO10': [0, 2, 3], 'PO11': [3, 1, 2],
            'PSO1': [2, 3, 1], 'PSO2': [1, 1, 2]
        }
        df_copom = pd.DataFrame(data, columns=all_cols)

        with pd.ExcelWriter(cb, engine='openpyxl') as writer:
            sheet_name = "CO_PO_Mapping"
            df_copom.to_excel(writer, sheet_name=sheet_name, index=False, startrow=5)
            write_metadata_to_sheet(writer, sheet_name, college_name, dept_name, course_code, len(df_copom.columns))
            
        cb.seek(0)
        z.writestr(f'{course_code}_co_po_mapping_sample.xlsx', cb.getvalue()) 

        # 3. Survey
        sb2 = BytesIO()
        df_survey = pd.DataFrame({
            "USN": ["1RV01", "1RV02", "1RV03"],
            "CO1_Rating": [3, 2, 3],
            "CO2_Rating": [2, 1, 2],
            "CO3_Rating": [3, 2, 3]
        })
        with pd.ExcelWriter(sb2, engine='openpyxl') as writer:
            sheet_name = "Survey"
            df_survey.to_excel(writer, sheet_name=sheet_name, index=False, startrow=5)
            write_metadata_to_sheet(writer, sheet_name, college_name, dept_name, course_code, len(df_survey.columns))
            
        sb2.seek(0)
        z.writestr(f'{course_code}_survey_data_sample.xlsx', sb2.getvalue()) 

    mem_zip.seek(0)
    return send_file(mem_zip, as_attachment=True,
                     download_name=f'{course_code}_all_samples.zip', 
                     mimetype='application/zip')


if __name__ == '__main__':
    for rule in app.url_map.iter_rules():
        print(rule)
    app.run(debug=True, port=5001)
